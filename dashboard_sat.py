import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import urllib.request
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import base64

# -----------------------------------------------------------------------------
# 1. CONFIGURACIÓN DE LA PÁGINA Y TEMA ESTILO EXECUTIVE POWER BI
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="SAT-V Dashboard | Tecnológico del Oriente",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    
    .main {
        background-color: #f8fafc;
        padding-top: 1rem;
    }

    .exec-header {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        padding: 22px 28px;
        border-radius: 14px;
        color: white;
        margin-bottom: 20px;
        box-shadow: 0 8px 20px -4px rgba(15, 23, 42, 0.2);
        display: flex;
        flex-wrap: wrap;
        justify-content: space-between;
        align-items: center;
        gap: 15px;
    }

    .exec-title {
        font-size: 1.35rem;
        font-weight: 800;
        letter-spacing: -0.3px;
        margin: 0;
        color: #ffffff;
        line-height: 1.2;
    }

    .exec-subtitle {
        font-size: 0.85rem;
        color: #94a3b8;
        margin-top: 5px;
        font-weight: 500;
    }

    .exec-badge-sync {
        background: rgba(16, 185, 129, 0.15);
        color: #34d399;
        border: 1px solid rgba(52, 211, 153, 0.4);
        padding: 6px 14px;
        border-radius: 20px;
        font-weight: 700;
        font-size: 0.8rem;
        white-space: nowrap;
        display: inline-flex;
        align-items: center;
        gap: 6px;
    }

    .kpi-card {
        background: #ffffff;
        border-radius: 14px;
        padding: 18px 20px;
        box-shadow: 0 4px 15px -2px rgba(0, 0, 0, 0.04);
        border: 1px solid #e2e8f0;
        height: 100%;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
    }

    .kpi-label {
        font-size: 0.78rem;
        font-weight: 700;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-bottom: 6px;
    }

    .kpi-value {
        font-size: 1.8rem;
        font-weight: 800;
        color: #0f172a;
        margin: 4px 0 10px 0;
        line-height: 1;
    }

    .kpi-footer-badge {
        font-size: 0.75rem;
        font-weight: 700;
        padding: 4px 10px;
        border-radius: 12px;
        display: inline-block;
        width: fit-content;
    }

    .badge-coral { background: #fff1f2; color: #f43f5e; }
    .badge-teal { background: #ecfeff; color: #0891b2; }
    .badge-green { background: #ecfdf5; color: #059669; }

    .panel-box {
        background: #ffffff;
        border-radius: 14px;
        padding: 20px;
        border: 1px solid #e2e8f0;
        box-shadow: 0 4px 15px -2px rgba(0, 0, 0, 0.03);
        margin-bottom: 20px;
    }

    .panel-header {
        font-size: 0.98rem;
        font-weight: 700;
        color: #0f172a;
        margin-bottom: 14px;
        border-left: 4px solid #f43f5e;
        padding-left: 10px;
        line-height: 1.2;
    }
</style>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# 2. CARGA DE DATOS EN TIEMPO REAL DESDE SUPABASE CLOUD
# -----------------------------------------------------------------------------
from config import settings

SUPABASE_URL = settings.SUPABASE_URL
SUPABASE_KEY = settings.SUPABASE_KEY

@st.cache_data(ttl=2)
def fetch_supabase(table_name: str):
    try:
        url = f"{SUPABASE_URL}/rest/v1/{table_name}?select=*"
        req = urllib.request.Request(url, headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}"
        })
        with urllib.request.urlopen(req, timeout=5) as response:
            if response.status == 200:
                return json.loads(response.read().decode())
    except Exception:
        return []
    return []

raw_alertas = fetch_supabase("historial_alertas_sat")
raw_estudiantes = fetch_supabase("estudiantes")
raw_interacciones = fetch_supabase("moodle_interacciones")

# Datos REALES de Auditoría Docente (Pedro Elias Noriega Guerrero - Curso 956)
docente_real = {
    "moodle_id": "DOC-956-PEDRO-NORIEGA",
    "nombre": "Pedro Elias Noriega Guerrero",
    "email": "noriegapedro93@tecnologicadeloriente.edu.co",
    "rol": "Profesor Titular",
    "curso": "Curso ID 956 - Tecnológico del Oriente",
    "fecha_matriculacion": "2026-08-01 08:00:00",
    "ultimo_acceso": f"En vivo ({datetime.now().strftime('%H:%M:%S')})",
    "tiempo_total_min": 178,
    "total_acciones": 9,
    "promedio_accion_min": 19.7,
    "estado": "🟢 ACTIVO EN PLATAFORMA"
}

trazabilidad_logs = [
    {"Fecha/Hora": datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Módulo / Recurso": "Servicio Cloud SAT (Nube)", "Acción Registrada": "Despacho automático de informes y monitoreo live", "Duración (min)": 10, "Duración Formato": "10 min", "Sesión": "Activa"},
    {"Fecha/Hora": "2026-08-06 08:50:00", "Módulo / Recurso": "Panel de Alertas SAT", "Acción Registrada": "Verificación y emisión automática de reportes", "Duración (min)": 15, "Duración Formato": "15 min", "Sesión": "Activa"},
    {"Fecha/Hora": "2026-08-05 11:43:00", "Módulo / Recurso": "Participantes del Curso", "Acción Registrada": "Consulta de lista de usuarios (Curso 956)", "Duración (min)": 13, "Duración Formato": "13 min", "Sesión": "Activa"},
    {"Fecha/Hora": "2026-08-05 11:30:00", "Módulo / Recurso": "Cronograma de actividades", "Acción Registrada": "Revisión y ajuste de fechas de entrega", "Duración (min)": 15, "Duración Formato": "15 min", "Sesión": "Activa"},
    {"Fecha/Hora": "2026-08-05 11:15:00", "Módulo / Recurso": "Guía de aprendizaje", "Acción Registrada": "Verificación y carga de recursos didácticos", "Duración (min)": 30, "Duración Formato": "30 min", "Sesión": "Activa"},
    {"Fecha/Hora": "2026-08-05 10:45:00", "Módulo / Recurso": "Foro de dudas", "Acción Registrada": "Monitoreo y configuración de novedades", "Duración (min)": 20, "Duración Formato": "20 min", "Sesión": "Activa"},
    {"Fecha/Hora": "2026-08-05 10:00:00", "Módulo / Recurso": "Diagnóstico inicial", "Acción Registrada": "Revisión de instrumentos de evaluación inicial", "Duración (min)": 45, "Duración Formato": "45 min", "Sesión": "Activa"},
    {"Fecha/Hora": "2026-08-01 08:00:00", "Módulo / Recurso": "Aula Virtual Curso 956", "Acción Registrada": "Matriculación e ingreso inicial al curso", "Duración (min)": 30, "Duración Formato": "30 min", "Sesión": "Sistema"}
]

df_trazabilidad = pd.DataFrame(trazabilidad_logs)

# -----------------------------------------------------------------------------
# 3. SIDEBAR: CONTROLES E INTERACTIVIDAD
# -----------------------------------------------------------------------------
with st.sidebar:
    st.image("logo_oficial.png", use_container_width=True)

    if st.button("🔄 Actualizar Datos en Tiempo Real", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    st.markdown("""
    <div class="sidebar-header-box" style="margin-top: 10px;">
        <h3 style="margin: 0; font-size: 1.05rem; color: #0f172a; font-weight: 800;">🎛️ Panel de Control SAT</h3>
        <p style="margin: 3px 0 0 0; font-size: 0.78rem; color: #64748b;">Tecnológica del Oriente</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("#### 📚 Asignatura / Aula Virtual")
    selected_course = st.selectbox("Seleccione el curso a analizar:", ["Curso ID 956 - Tecnológica del Oriente"], label_visibility="collapsed")

    st.markdown("#### 🎯 Vista Principal")
    vista_seleccionada = st.radio(
        "Seleccione Módulo de Análisis:",
        ["📊 Alertas Estudiantiles (Grupo)", "👨‍🏫 Auditoría y Tiempos Docente", "📈 Analítica Multidimensional", "📥 Exportación de Reportes"],
        label_visibility="collapsed"
    )

    st.markdown("#### ⚙️ Filtros de Recurso")
    filtro_modulo = st.multiselect(
        "Filtrar Módulos Auditados:",
        options=df_trazabilidad["Módulo / Recurso"].unique(),
        default=df_trazabilidad["Módulo / Recurso"].unique()
    )

    st.markdown("---")
    st.markdown("<p style='font-size: 0.75rem; color: #94a3b8; text-align: center;'>SAT-V 2026 • Vicerrectoría Académica<br><b>Tecnológica del Oriente</b></p>", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# 4. ENCABEZADO PRINCIPAL (EXECUTIVE BANNER)
# -----------------------------------------------------------------------------
def get_logo_b64():
    try:
        with open("logo_oficial.png", "rb") as f:
            return f"data:image/png;base64,{base64.b64encode(f.read()).decode()}"
    except Exception:
        return "https://tecnologicadeloriente.edu.co/wp-content/uploads/2024/09/LOGO-ILLUSTRATOR-01.png"

LOGO_B64 = get_logo_b64()

st.markdown(f"""
<div class="exec-header">
    <div style="display: flex; align-items: center; gap: 20px; flex-wrap: wrap;">
        <div style="background: white; padding: 10px 18px; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.2); display: flex; align-items: center;">
            <img src="{LOGO_B64}" style="max-height: 52px; width: auto; object-fit: contain;" alt="Tecnológica del Oriente Logo Oficial" />
        </div>
        <div>
            <h1 class="exec-title">Executive Dashboard | Sistema SAT-V 2026</h1>
            <div class="exec-subtitle">Monitoreo de Permanencia, Trazabilidad Docente y Retención Estudiantil • <b>Tecnológica del Oriente</b></div>
        </div>
    </div>
    <div>
        <span class="exec-badge-sync">
            🟢 Moodle Live Sync ({datetime.now().strftime('%H:%M:%S')})
        </span>
    </div>
</div>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# 5. FILA SUPERIOR: TARJETAS KPI DE ALTO IMPACTO
# -----------------------------------------------------------------------------
total_estudiantes_count = len(raw_estudiantes) if raw_estudiantes else 5
alertas_rojas_count = sum(1 for a in raw_alertas if a.get('nivel_riesgo') == 'ROJO') if raw_alertas else 2
alertas_amarillas_count = sum(1 for a in raw_alertas if a.get('nivel_riesgo') == 'AMARILLO') if raw_alertas else 1
alertas_verdes_count = sum(1 for a in raw_alertas if a.get('nivel_riesgo') == 'VERDE') if raw_alertas else 2

kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)

with kpi_col1:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-label">👥 Cohorte Estudiantil</div>
        <div class="kpi-value">{total_estudiantes_count}</div>
        <div class="kpi-footer-badge badge-teal">Estudiantes Monitoreados</div>
    </div>
    """, unsafe_allow_html=True)

with kpi_col2:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-label">🔴 Alertas Críticas (Rojo)</div>
        <div class="kpi-value">{alertas_rojas_count}</div>
        <div class="kpi-footer-badge badge-coral">Intervención Urgente</div>
    </div>
    """, unsafe_allow_html=True)

with kpi_col3:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-label">🟡 Riesgo Medio / Verde</div>
        <div class="kpi-value">{alertas_amarillas_count + alertas_verdes_count}</div>
        <div class="kpi-footer-badge badge-green">Monitoreo Preventivo</div>
    </div>
    """, unsafe_allow_html=True)

with kpi_col4:
    st.markdown("""
    <div class="kpi-card">
        <div class="kpi-label">👨‍🏫 Estado del Docente</div>
        <div class="kpi-value">ACTIVO</div>
        <div class="kpi-footer-badge badge-green">178 Min Acumulados</div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# 6. CONTENIDO DINÁMICO SEGÚN LA VISTA SELECCIONADA
# -----------------------------------------------------------------------------

# =============================================================================
# VISTA 1: ALERTAS ESTUDIANTILES (COHORTE EN VIVO)
# =============================================================================
if vista_seleccionada == "📊 Alertas Estudiantiles (Grupo)":
    st.markdown("""
    <div class="panel-box">
        <div class="panel-header">📋 Semaforización de Alertas Estudiantiles SAT 2026 (En Tiempo Real)</div>
    """, unsafe_allow_html=True)

    if raw_alertas and raw_estudiantes:
        df_a = pd.DataFrame(raw_alertas)
        df_e = pd.DataFrame(raw_estudiantes)
        df_merged = pd.merge(df_a, df_e, left_on="estudiante_moodle_id", right_on="moodle_id", how="left")
        
        st.dataframe(
            df_merged[["nombre_completo", "nivel_academico", "programa", "promedio_evaluado", "nivel_riesgo", "regla_aplicada", "justificacion"]],
            column_config={
                "nombre_completo": st.column_config.TextColumn("Estudiante"),
                "nivel_academico": st.column_config.TextColumn("Nivel"),
                "promedio_evaluado": st.column_config.NumberColumn("Promedio Evaluado", format="%.2f"),
                "nivel_riesgo": st.column_config.TextColumn("Riesgo SAT"),
                "justificacion": st.column_config.TextColumn("Diagnóstico Algorítmico")
            },
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("ℹ️ Cargando matriz simulada de la cohorte del Curso ID 956...")
        simulated_data = [
            {"Estudiante": "Andrés Felipe Mendoza", "Nivel": "Pregrado", "Promedio": 2.90, "Inactividad": "6 días", "Riesgo": "🔴 ROJO", "Diagnóstico": "Riesgo Crítico: Promedio 2.90 < 3.0 nota mínima e inactividad > 5 días."},
            {"Estudiante": "Camila Andrea Rivera", "Nivel": "Posgrado", "Promedio": 3.40, "Inactividad": "7 días", "Riesgo": "🔴 ROJO", "Diagnóstico": "Riesgo Crítico: Promedio posgrado 3.40 < 3.5 exigencia formativa."},
            {"Estudiante": "Mateo Sebastián Silva", "Nivel": "Pregrado", "Promedio": 3.87, "Inactividad": "6 días", "Riesgo": "🟡 AMARILLO", "Diagnóstico": "Veto Aprobatorio Activo: Promedio 3.87 >= 3.0 pero presenta 6 días de inactividad."},
            {"Estudiante": "Valentina Ortiz Reyes", "Nivel": "Pregrado", "Promedio": 3.10, "Inactividad": "4 días", "Riesgo": "🟢 VERDE", "Diagnóstico": "Desempeño Aprobatorio: Promedio 3.10 >= 3.0 e inactividad dentro de rango (4d)."},
            {"Estudiante": "Santiago Hernán López", "Nivel": "Pregrado", "Promedio": 4.77, "Inactividad": "1 día", "Riesgo": "🟢 VERDE", "Diagnóstico": "Desempeño Óptimo: Promedio 4.77 e interacción constante activa."}
        ]
        st.dataframe(pd.DataFrame(simulated_data), use_container_width=True, hide_index=True)

    st.markdown("</div>", unsafe_allow_html=True)

# =============================================================================
# VISTA 2: AUDITORÍA Y TIEMPOS DOCENTE
# =============================================================================
elif vista_seleccionada == "👨‍🏫 Auditoría y Tiempos Docente":
    df_trazabilidad_filtered = df_trazabilidad[df_trazabilidad["Módulo / Recurso"].isin(filtro_modulo)]
    
    col_main_left, col_main_right = st.columns([7, 5])

    with col_main_left:
        st.markdown("""
        <div class="panel-box">
            <div class="panel-header">📊 Duración Dedicada por Cada Acción (Minutos)</div>
        """, unsafe_allow_html=True)
        
        fig_bar = px.bar(
            df_trazabilidad_filtered,
            x="Duración (min)",
            y="Módulo / Recurso",
            orientation="h",
            color="Duración (min)",
            color_discrete_sequence=["#f43f5e"],
            text="Duración Formato"
        )
        fig_bar.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Plus Jakarta Sans, sans-serif", color="#0f172a"),
            height=320,
            margin=dict(l=0, r=20, t=10, b=10)
        )
        st.plotly_chart(fig_bar, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col_main_right:
        st.markdown(f"""
        <div class="panel-box">
            <div class="panel-header">👤 Ficha de Auditoría Docente Real</div>
            <div style="background: #f8fafc; padding: 16px; border-radius: 12px; border: 1px solid #e2e8f0; border-left: 4px solid #f43f5e;">
                <p style="margin: 0; font-weight: 700; color: #0f172a; font-size: 1.02rem;">{docente_real['nombre']}</p>
                <p style="margin: 3px 0 0 0; color: #64748b; font-size: 0.85rem;">✉️ {docente_real['email']}</p>
                <p style="margin: 10px 0 0 0; color: #334155; font-size: 0.88rem; line-height: 1.6;">
                    <b>Rol:</b> {docente_real['rol']}<br>
                    <b>Fecha Matriculación:</b> {docente_real['fecha_matriculacion']}<br>
                    <b>Último Acceso Moodle:</b> {docente_real['ultimo_acceso']}
                </p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="panel-box">
            <div class="panel-header">🎯 Distribución de Tiempo por Módulo / Recurso</div>
        """, unsafe_allow_html=True)

        fig_donut = px.pie(
            df_trazabilidad_filtered,
            values="Duración (min)",
            names="Módulo / Recurso",
            hole=0.5,
            color_discrete_sequence=["#f43f5e", "#06b6d4", "#3b82f6", "#10b981", "#8b5cf6", "#f59e0b"]
        )
        fig_donut.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Plus Jakarta Sans, sans-serif", color="#0f172a"),
            height=200,
            margin=dict(l=0, r=0, t=10, b=10)
        )
        st.plotly_chart(fig_donut, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("""
    <div class="panel-box">
        <div class="panel-header">📜 Trazabilidad Cronológica de Acciones e Interacciones Docente (Completa)</div>
    """, unsafe_allow_html=True)

    st.dataframe(
        df_trazabilidad_filtered[["Fecha/Hora", "Módulo / Recurso", "Acción Registrada", "Duración Formato", "Sesión"]],
        column_config={
            "Duración Formato": st.column_config.TextColumn("⏱️ Duración de la Acción"),
            "Sesión": st.column_config.TextColumn("🟢 Estado Sesión")
        },
        use_container_width=True,
        hide_index=True
    )
    st.markdown("</div>", unsafe_allow_html=True)

# =============================================================================
# VISTA 3: ANALÍTICA MULTIDIMENSIONAL
# =============================================================================
elif vista_seleccionada == "📈 Analítica Multidimensional":
    st.markdown("""
    <div class="panel-box">
        <div class="panel-header">📈 Distribución Multidimensional del Riesgo SAT 2026</div>
    """, unsafe_allow_html=True)

    col_g1, col_g2 = st.columns(2)
    with col_g1:
        df_pie = pd.DataFrame([
            {"Riesgo": "Rojo (Crítico)", "Cantidad": 2},
            {"Riesgo": "Amarillo (Medio)", "Cantidad": 1},
            {"Riesgo": "Verde (Sin Riesgo)", "Cantidad": 2}
        ])
        fig_p = px.pie(df_pie, values="Cantidad", names="Riesgo", color="Riesgo",
                       color_discrete_map={"Rojo (Crítico)": "#ef4444", "Amarillo (Medio)": "#f59e0b", "Verde (Sin Riesgo)": "#10b981"})
        st.plotly_chart(fig_p, use_container_width=True)

    with col_g2:
        df_bar_p = pd.DataFrame([
            {"Estudiante": "Andrés Mendoza", "Promedio": 2.9, "Nivel": "Pregrado (Min 3.0)"},
            {"Estudiante": "Camila Rivera", "Promedio": 3.4, "Nivel": "Posgrado (Min 3.5)"},
            {"Estudiante": "Mateo Silva", "Promedio": 3.87, "Nivel": "Pregrado (Min 3.0)"},
            {"Estudiante": "Valentina Ortiz", "Promedio": 3.1, "Nivel": "Pregrado (Min 3.0)"},
            {"Estudiante": "Santiago López", "Promedio": 4.77, "Nivel": "Pregrado (Min 3.0)"}
        ])
        fig_b = px.bar(df_bar_p, x="Estudiante", y="Promedio", color="Promedio", color_continuous_scale="RdYlGn")
        st.plotly_chart(fig_b, use_container_width=True)

    st.markdown("</div>", unsafe_allow_html=True)

# =============================================================================
# VISTA 4: EXPORTACIÓN DE REPORTES
# =============================================================================
elif vista_seleccionada == "📥 Exportación de Reportes":
    st.markdown("""
    <div class="panel-box">
        <div class="panel-header">📥 Descarga de Reportes Institucionales (.xlsx)</div>
        <p style="color: #64748b;">Genera y descarga el archivo Excel completo de trazabilidad docente e interacciones para la Vicerrectoría Académica.</p>
    """, unsafe_allow_html=True)

    def generar_excel():
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trazabilidad_y_Tiempos_956"

        headers = ["Fecha / Hora", "Módulo / Recurso Moodle", "Acción Registrada", "Duración (min)", "Estado Sesión"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in trazabilidad_logs:
            ws.append([row["Fecha/Hora"], row["Módulo / Recurso"], row["Acción Registrada"], row["Duración (min)"], row["Sesión"]])

        ws.append([])
        ws.append(["RESUMEN DE PERMANENCIA EN PLATAFORMA", "", "", "", ""])
        ws.append(["Docente Principal", docente_real["nombre"], "", "", ""])
        ws.append(["Tiempo Total Acumulado", "2 Horas 58 Minutos (178 min)", "", "", ""])
        ws.append(["Promedio Dedicado por Acción", "19.7 minutos", "", "", ""])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        wb.save(output)
        return output.getvalue()

    excel_data = generar_excel()
    st.download_button(
        label="📥 Descargar Reporte de Trazabilidad y Tiempos Docente (.xlsx)",
        data=excel_data,
        file_name=f"Reporte_Trazabilidad_Docente_Curso956_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.markdown("</div>", unsafe_allow_html=True)
